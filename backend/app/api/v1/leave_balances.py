"""Leave balances: opening entry, ledger, projection, annual credit, carry-forward, manual adjust."""

import io
import json
import uuid
from datetime import date

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth.dependencies import employee_scope, get_current_user, require_role
from app.core.cache import cache_clear, cached, cache_set
from app.core.database import get_db
from app.core.upload_validation import validate_import_upload
from app.services.leave_ledger import record_leave_ledger

router = APIRouter(prefix="/leave-balances", tags=["leave-balances"])

_ADMIN_BALANCE_ROLES = ("ADMIN", "ESTABLISHMENT_OFFICER")
_NODAL_ADJUST_ROLES = ("NODAL_OFFICER",)
_ADJUSTABLE_FIELDS = {"opening_balance", "credited", "availed", "lop_days"}


def _projection_cache_key(employee_id: str, leave_type_code: str, from_date: str, to_date: str) -> str:
    return f"leave_projection:{employee_id}:{leave_type_code}:{from_date}:{to_date}"


def _employee_in_scope(scope: dict, employee_id: str) -> bool:
    employee_ids = scope.get("employee_ids")
    return employee_ids is None or employee_id in employee_ids


def _require_employee_scope(scope: dict, employee_id: str) -> None:
    if not _employee_in_scope(scope, employee_id):
        raise HTTPException(status_code=403, detail="Employee outside permitted scope")


def _parse_adjustment_payload(raw: str | None) -> dict:
    if not raw:
        return {}
    if isinstance(raw, dict):
        return raw
    try:
        return json.loads(raw)
    except Exception:
        return {}


def _working_days(from_date: date, to_date: date, holidays: set[date]) -> int:
    return sum(
        1
        for d in [from_date + date.resolution * i for i in range((to_date - from_date).days + 1)]
        if d.weekday() < 5 and d not in holidays
    )


@router.post("/opening")
async def bulk_opening_balance(
    body: list[dict],
    current_user: dict = Depends(require_role(*_ADMIN_BALANCE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """Accepts array of {emp_code, leave_type_code, opening_balance}. Idempotent."""
    results = []
    for item in body:
        emp_code = item["emp_code"]
        lt_code = item["leave_type_code"]
        balance = item["opening_balance"]

        emp = await db.execute(
            text("""
                SELECT e.id, c.leave_scheme, e.doj
                FROM employees e
                JOIN employee_categories c ON e.category_id = c.id
                WHERE e.emp_code = :ec
            """),
            {"ec": emp_code},
        )
        emp_row = emp.fetchone()
        if not emp_row:
            results.append({"emp_code": emp_code, "status": "error", "message": "Employee not found"})
            continue

        lt = await db.execute(text("SELECT id FROM leave_types WHERE code = :c"), {"c": lt_code})
        lt_row = lt.fetchone()
        if not lt_row:
            results.append({"emp_code": emp_code, "status": "error", "message": f"Unknown leave type: {lt_code}"})
            continue

        scheme = emp_row.leave_scheme
        leave_year = 2026
        if scheme == "CCS":
            year_start = date(leave_year, 4, 1)
        else:
            doj = emp_row.doj
            year_start = date(leave_year, doj.month, doj.day)

        existing = await db.execute(
            text("""
                SELECT id
                FROM leave_balances
                WHERE employee_id = :eid AND leave_type_id = :lid AND leave_year = :ly
            """),
            {"eid": str(emp_row.id), "lid": str(lt_row[0]), "ly": leave_year},
        )
        ex_row = existing.fetchone()

        if ex_row:
            bid = str(ex_row[0])
            await db.execute(
                text("""
                    UPDATE leave_balances
                    SET opening_balance = :bal, credited = 0, availed = 0, last_updated = now()
                    WHERE id = :id
                """),
                {"bal": balance, "id": bid},
            )
        else:
            bid = str(uuid.uuid4())
            await db.execute(
                text("""
                    INSERT INTO leave_balances (id, employee_id, leave_type_id, leave_year, year_start_date, opening_balance, credited)
                    VALUES (:id, :eid, :lid, :ly, :ysd, :bal, 0)
                """),
                {"id": bid, "eid": str(emp_row.id), "lid": str(lt_row[0]), "ly": leave_year, "ysd": year_start, "bal": balance},
            )
        await record_leave_ledger(
            db,
            balance_id=bid,
            employee_id=str(emp_row.id),
            leave_type_id=str(lt_row[0]),
            leave_year=leave_year,
            txn_type="OPENING",
            amount=float(balance),
            field_affected="opening_balance",
            reference_type="system",
            reason="Opening balance entry",
            actor_id=current_user["user_id"],
            impersonated_by=current_user.get("impersonated_by"),
        )
        results.append({"emp_code": emp_code, "leave_type": lt_code, "status": "ok"})

    await db.commit()
    cache_clear()
    return {"processed": len(results), "results": results}


@router.post("/opening/import")
async def import_opening_balances(
    file: UploadFile,
    _: dict = Depends(require_role(*_ADMIN_BALANCE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """Excel import. Columns: emp_code, leave_type_code, opening_balance."""
    await validate_import_upload(
        file,
        allowed_extensions={".xlsx"},
        allowed_content_types={"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        label="Opening balance XLSX import",
    )
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    payload = []
    for row in rows:
        if not row[0]:
            continue
        payload.append(
            {
                "emp_code": str(row[0]).strip(),
                "leave_type_code": str(row[1]).strip() if row[1] else "",
                "opening_balance": float(row[2]) if row[2] else 0,
            }
        )
    return await bulk_opening_balance(payload, _=_, db=db)


@router.get("/{employee_id}")
async def get_balances(
    employee_id: str,
    current_user: dict = Depends(get_current_user),
    scope: dict = Depends(employee_scope),
    db: AsyncSession = Depends(get_db),
):
    _require_employee_scope(scope, employee_id)
    result = await db.execute(
        text("""
            SELECT lb.*, lt.code AS leave_type_code, lt.name AS leave_type_name, lt.max_accumulation
            FROM leave_balances lb
            JOIN leave_types lt ON lb.leave_type_id = lt.id
            WHERE lb.employee_id = :eid
            ORDER BY lb.leave_year DESC, lt.code ASC
        """),
        {"eid": employee_id},
    )
    balances = [dict(r._mapping) for r in result.fetchall()]
    return {"employee_id": employee_id, "balances": balances}


@router.get("/{employee_id}/ledger/{leave_type_id}")
async def get_ledger(
    employee_id: str,
    leave_type_id: str,
    current_user: dict = Depends(get_current_user),
    scope: dict = Depends(employee_scope),
    db: AsyncSession = Depends(get_db),
):
    _require_employee_scope(scope, employee_id)

    bal = await db.execute(
        text("""
            SELECT lb.*, lt.code AS leave_type_code, lt.name AS leave_type_name, lt.max_accumulation
            FROM leave_balances lb
            JOIN leave_types lt ON lb.leave_type_id = lt.id
            WHERE lb.employee_id = :eid AND lb.leave_type_id = :lid
            ORDER BY lb.leave_year
        """),
        {"eid": employee_id, "lid": leave_type_id},
    )
    balances = [dict(r._mapping) for r in bal.fetchall()]

    apps = await db.execute(
        text("""
            SELECT app_number, from_date, to_date, applied_days, status, submitted_at, updated_at
            FROM leave_applications
            WHERE employee_id = :eid AND leave_type_id = :lid AND status IN ('APPROVED', 'RECALLED')
            ORDER BY COALESCE(updated_at, submitted_at), from_date
        """),
        {"eid": employee_id, "lid": leave_type_id},
    )
    application_events = [
        {
            "entry_type": "application",
            "event_date": row.updated_at or row.submitted_at,
            "app_number": row.app_number,
            "from_date": row.from_date,
            "to_date": row.to_date,
            "days": float(row.applied_days),
            "delta": -float(row.applied_days) if row.status == "APPROVED" else float(row.applied_days),
            "status": row.status,
        }
        for row in apps.fetchall()
    ]

    balance_ids = [row["id"] for row in balances]
    ledger_transactions = []
    if balance_ids:
        ledger_rows = await db.execute(
            text("""
                SELECT lbl.created_at, lbl.txn_type, lbl.amount, lbl.field_affected, lbl.reason,
                       lbl.reference_type, lbl.balance_id, lbl.leave_year, la.app_number
                FROM leave_balance_ledger lbl
                LEFT JOIN leave_applications la
                  ON lbl.reference_id = la.id AND lbl.reference_type = 'application'
                WHERE lbl.balance_id = ANY(:balance_ids)
                ORDER BY lbl.created_at
            """),
            {"balance_ids": balance_ids},
        )
        for row in ledger_rows.fetchall():
            field = row.field_affected
            amt = float(row.amount or 0)
            ledger_transactions.append(
                {
                    "entry_type": str(row.txn_type).lower(),
                    "event_date": row.created_at,
                    "balance_id": str(row.balance_id),
                    "leave_year": row.leave_year,
                    "field": field,
                    "reason": row.reason,
                    "days": abs(amt),
                    "delta": -amt if field == "availed" and row.txn_type == "AVAIL" else amt,
                    "app_number": row.app_number,
                }
            )

    manual_adjustments = []
    if balance_ids and not ledger_transactions:
        audit_rows = await db.execute(
            text("""
                SELECT entity_id, created_at, after_state
                FROM audit_log
                WHERE entity_type = 'leave_balance' AND entity_id = ANY(:balance_ids)
                ORDER BY created_at
            """),
            {"balance_ids": balance_ids},
        )
        for row in audit_rows.fetchall():
            payload = _parse_adjustment_payload(row.after_state)
            if payload:
                amount = float(payload.get("amount", 0))
                field = str(payload.get("field", "credited"))
                manual_adjustments.append(
                    {
                        "entry_type": "manual_adjustment",
                        "event_date": row.created_at,
                        "balance_id": str(row.entity_id),
                        "field": field,
                        "reason": payload.get("reason"),
                        "days": amount,
                        "delta": amount if field != "availed" else -amount,
                    }
                )

    yearly_events = []
    if not ledger_transactions:
        for row in balances:
            yearly_events.append(
                {
                    "entry_type": "opening_balance",
                    "event_date": row["year_start_date"],
                    "balance_id": str(row["id"]),
                    "leave_year": row["leave_year"],
                    "days": float(row["opening_balance"] or 0),
                    "delta": float(row["opening_balance"] or 0),
                }
            )
            credited = float(row["credited"] or 0)
            if credited:
                yearly_events.append(
                    {
                        "entry_type": "annual_credit",
                        "event_date": row["year_start_date"],
                        "balance_id": str(row["id"]),
                        "leave_year": row["leave_year"],
                        "days": credited,
                        "delta": credited,
                    }
                )

    emp_res = await db.execute(text("SELECT category_id FROM employees WHERE id = :eid"), {"eid": employee_id})
    emp_row = emp_res.fetchone()
    cat_id = emp_row.category_id if emp_row else None

    wc_res = await db.execute(
        text("""
            SELECT id, config_name, min_days, max_days 
            FROM workflow_configs 
            WHERE (leave_type_id = :lid OR leave_type_id IS NULL)
              AND is_active = true
              AND (category_id = :cat OR category_id IS NULL)
            ORDER BY 
                 (CASE WHEN category_id IS NOT NULL THEN 1 ELSE 0 END) DESC,
                 (CASE WHEN leave_type_id IS NOT NULL THEN 1 ELSE 0 END) DESC
            LIMIT 1
        """),
        {"lid": leave_type_id, "cat": cat_id}
    )
    approval_chains = []
    for r in wc_res.fetchall():
        cfg = dict(r._mapping)
        steps_res = await db.execute(
            text("SELECT step_order, approver_role FROM workflow_steps WHERE config_id = :cid ORDER BY step_order"),
            {"cid": cfg["id"]}
        )
        cfg["steps"] = [s.approver_role for s in steps_res.fetchall()]
        approval_chains.append(cfg)

    if ledger_transactions:
        transactions = ledger_transactions
    else:
        transactions = sorted(
            [*yearly_events, *manual_adjustments, *application_events],
            key=lambda item: (str(item.get("event_date") or ""), item["entry_type"]),
        )
    return {
        "employee_id": employee_id,
        "leave_type_id": leave_type_id,
        "balances": balances,
        "transactions": transactions,
        "approval_chains": approval_chains,
        "ledger_source": "immutable" if ledger_transactions else "legacy",
    }


@router.get("/{employee_id}/project")
async def project_balance(
    employee_id: str,
    from_date: str = Query(),
    to_date: str = Query(),
    leave_type_code: str = Query(),
    current_user: dict = Depends(get_current_user),
    scope: dict = Depends(employee_scope),
    db: AsyncSession = Depends(get_db),
):
    _require_employee_scope(scope, employee_id)
    cache_key = _projection_cache_key(employee_id, leave_type_code, from_date, to_date)
    cached_projection = cached(cache_key)
    if cached_projection is not None:
        return {**cached_projection, "cache_ttl_seconds": 300, "cached": True}

    lt = await db.execute(text("SELECT id FROM leave_types WHERE code = :c"), {"c": leave_type_code})
    lt_row = lt.fetchone()
    if not lt_row:
        raise HTTPException(status_code=400, detail=f"Unknown leave type: {leave_type_code}")

    lt_id = str(lt_row[0])
    bal = await db.execute(
        text("""
            SELECT closing_balance
            FROM leave_balances
            WHERE employee_id = :eid AND leave_type_id = :lid
            ORDER BY leave_year DESC
            LIMIT 1
        """),
        {"eid": employee_id, "lid": lt_id},
    )
    bal_row = bal.fetchone()
    current = float(bal_row[0]) if bal_row else 0

    fd = date.fromisoformat(from_date)
    td = date.fromisoformat(to_date)
    holidays_result = await db.execute(
        text("SELECT holiday_date FROM holiday_master WHERE holiday_date BETWEEN :f AND :t"),
        {"f": fd, "t": td},
    )
    holiday_dates = {r[0] for r in holidays_result.fetchall()}
    days = _working_days(fd, td, holiday_dates)
    payload = {
        "current_balance": current,
        "requested_days": days,
        "projected_balance": max(0, current - days),
    }
    cache_set(cache_key, payload)
    return {**payload, "cache_ttl_seconds": 300, "cached": False}


@router.post("/credit/annual")
async def annual_credit_run(
    body: dict,
    _: dict = Depends(require_role(*_ADMIN_BALANCE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    year_start = date.fromisoformat(body["year_start"])
    leave_year = body["leave_year"]
    updated = await db.execute(
        text("""
            UPDATE leave_balances lb
            SET credited = COALESCE(ler.days_per_year, 0),
                last_updated = now()
            FROM employees e
            JOIN employee_categories c ON e.category_id = c.id
            JOIN leave_entitlement_rules ler ON ler.category_id = c.id
            JOIN leave_types lt ON ler.leave_type_id = lt.id
            WHERE lb.employee_id = e.id
              AND lb.leave_type_id = lt.id
              AND lb.leave_year = :ly
              AND c.leave_scheme = 'CCS'
              AND ler.year_ref = 'FINANCIAL'
              AND lt.code IN ('EL', 'HPL')
              AND COALESCE(ler.days_per_year, 0) > 0
              AND COALESCE(lb.credited, 0) = 0
        """),
        {"ly": leave_year},
    )
    inserted = await db.execute(
        text("""
            INSERT INTO leave_balances (id, employee_id, leave_type_id, leave_year, year_start_date, opening_balance, credited)
            SELECT uuid_generate_v4(), e.id, lt.id, :ly, CAST(:ys AS date),
                   COALESCE((
                       SELECT closing_balance
                       FROM leave_balances
                       WHERE employee_id = e.id AND leave_type_id = lt.id AND leave_year = :ly - 1
                   ), 0),
                   COALESCE(ler.days_per_year, 0)
            FROM employees e
            JOIN employee_categories c ON e.category_id = c.id
            JOIN leave_entitlement_rules ler ON ler.category_id = c.id
            JOIN leave_types lt ON ler.leave_type_id = lt.id
            WHERE c.leave_scheme = 'CCS'
              AND ler.year_ref = 'FINANCIAL'
              AND lt.code IN ('EL', 'HPL')
              AND COALESCE(ler.days_per_year, 0) > 0
              AND NOT EXISTS (
                  SELECT 1
                  FROM leave_balances lb2
                  WHERE lb2.employee_id = e.id AND lb2.leave_type_id = lt.id AND lb2.leave_year = :ly
              )
        """),
        {"ly": leave_year, "ys": year_start},
    )
    await db.commit()
    cache_clear()
    rows_affected = (updated.rowcount or 0) + (inserted.rowcount or 0)
    return {"message": f"Annual credit run complete for {leave_year}", "rows_affected": rows_affected}


@router.post("/carryforward")
async def carry_forward_run(
    body: dict,
    _: dict = Depends(require_role(*_ADMIN_BALANCE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    source_year = body["source_year"]
    target_year = body["target_year"]
    year_start = date.fromisoformat(body["year_start"]) if body.get("year_start") else date(target_year, 4, 1)
    result = await db.execute(
        text("""
            INSERT INTO leave_balances (id, employee_id, leave_type_id, leave_year, year_start_date, opening_balance, credited)
            SELECT uuid_generate_v4(), lb.employee_id, lb.leave_type_id, :ty, CAST(:ys AS date),
                   LEAST(COALESCE(lt.max_accumulation, 300), lb.closing_balance),
                   0
            FROM leave_balances lb
            JOIN leave_types lt ON lb.leave_type_id = lt.id
            WHERE lb.leave_year = :sy
              AND lt.carry_forward = true
              AND NOT EXISTS (
                  SELECT 1
                  FROM leave_balances lb2
                  WHERE lb2.employee_id = lb.employee_id AND lb2.leave_type_id = lb.leave_type_id AND lb2.leave_year = :ty
              )
        """),
        {"sy": source_year, "ty": target_year, "ys": year_start},
    )
    await db.commit()
    cache_clear()
    return {"message": "Carry-forward complete", "rows_affected": result.rowcount}


@router.put("/{balance_id}/manual-adjust")
async def manual_adjust(
    balance_id: str,
    body: dict,
    current_user: dict = Depends(get_current_user),
    scope: dict = Depends(employee_scope),
    db: AsyncSession = Depends(get_db),
):
    role = current_user["role"]
    if role not in _ADMIN_BALANCE_ROLES and role not in _NODAL_ADJUST_ROLES:
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    reason = body.get("reason")
    if not reason:
        raise HTTPException(status_code=400, detail="Reason required")

    amount = float(body.get("amount", 0))
    field = body.get("field", "credited")
    if field not in _ADJUSTABLE_FIELDS:
        raise HTTPException(status_code=400, detail=f"Unsupported field: {field}")

    existing = await db.execute(
        text("""
            SELECT lb.id, lb.employee_id, lb.leave_type_id, lb.leave_year
            FROM leave_balances lb
            WHERE lb.id = :bid
        """),
        {"bid": balance_id},
    )
    bal_row = existing.fetchone()
    if not bal_row:
        raise HTTPException(status_code=404, detail="Leave balance not found")

    if role in _NODAL_ADJUST_ROLES:
        _require_employee_scope(scope, str(bal_row.employee_id))

    await db.execute(
        text(f"UPDATE leave_balances SET {field} = {field} + :amt, last_updated = now() WHERE id = :bid"),
        {"amt": amount, "bid": balance_id},
    )
    await record_leave_ledger(
        db,
        balance_id=balance_id,
        employee_id=str(bal_row.employee_id),
        leave_type_id=str(bal_row.leave_type_id),
        leave_year=int(bal_row.leave_year),
        txn_type="MANUAL_ADJUST",
        amount=amount,
        field_affected=field,
        reference_type="manual",
        reason=reason,
        actor_id=current_user["user_id"],
        impersonated_by=current_user.get("impersonated_by"),
    )
    await db.execute(
        text("""
            INSERT INTO audit_log (id, actor_id, impersonated_by, entity_type, entity_id, action, before_state, after_state)
            VALUES (uuid_generate_v4(), :aid, :impersonated_by, 'leave_balance', :bid, 'UPDATE', '{}', CAST(:reason AS jsonb))
        """),
        {
            "aid": current_user["user_id"],
            "impersonated_by": current_user.get("impersonated_by"),
            "bid": balance_id,
            "reason": json.dumps({"reason": reason, "field": field, "amount": amount}),
        },
    )
    await db.commit()
    cache_clear()
    return {"message": "Adjusted", "reason": reason}
