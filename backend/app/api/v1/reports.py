"""Reports routes -- locked Phase 7 report outputs and payroll export."""

import csv
import io
import json
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth.dependencies import require_role
from app.core.database import get_db
from app.core.rate_limit import limiter
from app.core.config import settings

router = APIRouter(prefix="/reports", tags=["reports"])

REPORT_ROLES = ("ESTABLISHMENT_OFFICER", "REGISTRAR", "DIRECTOR", "NODAL_OFFICER", "NODAL_OFFICE")
PAYROLL_EXPORT_ROLES = ("ESTABLISHMENT_OFFICER", "REGISTRAR", "DIRECTOR")

# Placeholder until AIIMS Finance / NIC provides the actual column contract.
PAYROLL_NIC_MAPPING_PLACEHOLDER = {
    "emp_code": "Emp Code",
    "name": "Name",
    "department": "Dept",
    "month": "Month",
    "lop_days": "LOP Days",
    "reason": "Reason",
}


_NODAL_SCOPED_ROLES = ("NODAL_OFFICER", "NODAL_OFFICE")


def _apply_nodal_scope(query: str, current_user: dict | None, params: dict, employee_alias: str = "e") -> str:
    if current_user and current_user.get("role") in _NODAL_SCOPED_ROLES:
        query += f" AND EXISTS (SELECT 1 FROM dept_nodal_assignments dna WHERE dna.department_id = {employee_alias}.department_id AND dna.nodal_user_id = :nodal_user_id)"
        params["nodal_user_id"] = current_user.get("user_id")
    return query


def _json_safe(value):
    if isinstance(value, Decimal):
        return float(value)
    return value


def _excel_safe(value):
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _rows_from_result(result) -> list[dict]:
    return [{key: _json_safe(value) for key, value in row._mapping.items()} for row in result.fetchall()]


def _workbook_bytes(sheet_name: str, headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append([_excel_safe(value) for value in row])

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", start_color="DCEBFA", end_color="DCEBFA")

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 40)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _pdf_bytes(title: str, headers: list[str], rows: list[list[object]]) -> bytes:
    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )

    table_data = [headers] + [[str(value if value is not None else "") for value in row] for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dce6f1")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f9fc")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Heading2"]), Spacer(1, 0.3 * cm), table]
    document.build(story)
    return buffer.getvalue()


def _stream_bytes(content: bytes, media_type: str, filename: str) -> StreamingResponse:
    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def _fetch_leave_register_rows(db: AsyncSession, from_date: date, to_date: date, department_code: str | None, current_user: dict | None = None) -> list[dict]:
    query = """
        SELECT
            e.emp_code AS emp_code,
            e.name AS name,
            d.name AS dept,
            lt.code AS leave_type,
            a.from_date AS from_date,
            a.to_date AS to_date,
            a.applied_days AS days,
            a.status AS status,
            (
                SELECT MAX(la.acted_at)
                FROM leave_approvals la
                JOIN workflow_steps ws ON ws.id = la.step_id
                WHERE la.application_id = a.id
                  AND la.action = 'APPROVED'
                  AND ws.is_final_authority = true
            ) AS approval_date
        FROM leave_applications a
        JOIN employees e ON a.employee_id = e.id
        JOIN departments d ON e.department_id = d.id
        JOIN leave_types lt ON a.leave_type_id = lt.id
        WHERE a.from_date <= :to_date AND a.to_date >= :from_date
    """
    params: dict[str, object] = {"from_date": from_date, "to_date": to_date}
    if department_code:
        query += " AND d.code = :department_code"
        params["department_code"] = department_code
    query = _apply_nodal_scope(query, current_user, params)
    query += " ORDER BY e.emp_code, a.from_date"
    result = await db.execute(text(query), params)
    return _rows_from_result(result)


async def _fetch_category_summary_rows(db: AsyncSession, from_date: date, to_date: date, current_user: dict | None = None) -> list[dict]:
    staff_result = await db.execute(
        text(
            "SELECT c.code AS category, COUNT(e.id) AS total_staff "
            "FROM employee_categories c "
            "LEFT JOIN employees e ON e.category_id = c.id AND e.is_active = true "
            + _apply_nodal_scope("", current_user, {}, "e") + 
            " GROUP BY c.code "
            "ORDER BY c.code"
        ), {"nodal_user_id": current_user.get("user_id")} if current_user and current_user.get("role") in _NODAL_SCOPED_ROLES else {}
    )
    staff_counts = {row._mapping["category"]: int(row._mapping["total_staff"] or 0) for row in staff_result.fetchall()}

    leave_result = await db.execute(
        text(
            "SELECT "
                "c.code AS category, "
                "lt.code AS leave_type, "
                "COALESCE(SUM(a.applied_days), 0) AS total_leave_days "
            "FROM employee_categories c "
            "LEFT JOIN employees e ON e.category_id = c.id "
            "LEFT JOIN leave_applications a "
              "ON a.employee_id = e.id "
             "AND a.status = 'APPROVED' "
             "AND a.from_date BETWEEN :from_date AND :to_date "
            "LEFT JOIN leave_types lt ON lt.id = a.leave_type_id "
            + _apply_nodal_scope("", current_user, {}, "e") + 
            " GROUP BY c.code, lt.code "
            "ORDER BY c.code, lt.code"
        ),
        {"from_date": from_date, "to_date": to_date, **({"nodal_user_id": current_user.get("user_id")} if current_user and current_user.get("role") in _NODAL_SCOPED_ROLES else {})},
    )

    grouped: dict[str, dict[str, object]] = {}
    by_type: dict[str, list[str]] = defaultdict(list)
    total_days_by_category: dict[str, float] = defaultdict(float)

    for row in leave_result.fetchall():
        category = row._mapping["category"]
        leave_type = row._mapping["leave_type"]
        total_leave_days = float(row._mapping["total_leave_days"] or 0)
        grouped.setdefault(category, {"category": category})
        if leave_type:
            by_type[category].append(f"{leave_type}: {total_leave_days:g}")
            total_days_by_category[category] += total_leave_days

    rows: list[dict] = []
    for category in sorted(staff_counts.keys()):
        total_staff = staff_counts[category]
        total_leave_days = total_days_by_category.get(category, 0.0)
        avg = round(total_leave_days / total_staff, 2) if total_staff else 0.0
        rows.append(
            {
                "category": category,
                "total_staff": total_staff,
                "total_leave_days_by_type": ", ".join(by_type.get(category, [])) or "No approved leave in range",
                "avg_per_staff": avg,
            }
        )
    return rows


async def _fetch_pending_rows(db: AsyncSession) -> list[dict]:
    result = await db.execute(
        text(
            """
            SELECT
                a.app_number AS app_number,
                e.name AS employee,
                lt.code AS leave_type,
                a.submitted_at AS submitted_date,
                FLOOR(EXTRACT(EPOCH FROM (now() - a.submitted_at)) / 86400) AS days_pending,
                COALESCE(
                    specific_user.username,
                    ws.approver_role
                ) AS current_approver
            FROM leave_applications a
            JOIN employees e ON a.employee_id = e.id
            JOIN leave_types lt ON a.leave_type_id = lt.id
            JOIN workflow_steps ws
              ON ws.config_id = a.config_id
             AND ws.step_order = a.current_step_order
            LEFT JOIN users specific_user ON specific_user.id = ws.specific_approver_id
            WHERE a.status IN ('SUBMITTED', 'UNDER_REVIEW')
            ORDER BY a.submitted_at
            """
        )
    )
    return _rows_from_result(result)


async def _fetch_balance_rows(
    db: AsyncSession,
    as_of_date: date | None,
    department_code: str | None = None,
    designation_name: str | None = None,
    current_user: dict | None = None,
) -> list[dict]:
    target_year = as_of_date.year if as_of_date else date.today().year
    query = """
            SELECT
                e.emp_code AS emp_code,
                e.name AS name,
                d.code AS department_code,
                d.name AS dept,
                des.name AS designation_name,
                lt.code AS leave_type,
                lb.opening_balance AS opening_balance,
                lb.credited AS credited,
                lb.availed AS availed,
                lb.closing_balance AS closing_balance
            FROM leave_balances lb
            JOIN employees e ON lb.employee_id = e.id
            JOIN departments d ON e.department_id = d.id
            JOIN designations des ON e.designation_id = des.id
            JOIN leave_types lt ON lb.leave_type_id = lt.id
            WHERE lb.leave_year = :leave_year AND e.is_active = true
    """
    params: dict = {"leave_year": target_year}
    if department_code and department_code != "ALL":
        query += " AND d.code = :dept_code"
        params["dept_code"] = department_code
    if designation_name and designation_name != "ALL":
        query += " AND des.name = :desg_name"
        params["desg_name"] = designation_name
    query = _apply_nodal_scope(query, current_user, params)
    query += " ORDER BY d.name, des.name, e.emp_code, lt.code"
    result = await db.execute(text(query), params)
    return _rows_from_result(result)


async def _fetch_calendar_rows(db: AsyncSession, department_code: str | None, month: str | None) -> list[dict]:
    query = """
        SELECT e.emp_code, e.name, a.from_date, a.to_date, lt.code AS leave_type
        FROM leave_applications a
        JOIN employees e ON a.employee_id = e.id
        JOIN departments d ON e.department_id = d.id
        JOIN leave_types lt ON a.leave_type_id = lt.id
        WHERE a.status = 'APPROVED'
    """
    params: dict[str, object] = {}
    if department_code:
        query += " AND d.code = :department_code"
        params["department_code"] = department_code
    if month:
        month_start = date.fromisoformat(f"{month}-01")
        month_end = date(month_start.year + (1 if month_start.month == 12 else 0), 1 if month_start.month == 12 else month_start.month + 1, 1)
        query += " AND a.from_date >= :month_start AND a.from_date < :month_end"
        params["month_start"] = month_start
        params["month_end"] = month_end
    query += " ORDER BY a.from_date, e.emp_code"
    result = await db.execute(text(query), params)
    return _rows_from_result(result)


@router.get("/leave-register")
@limiter.limit(settings.RATE_LIMIT_EXPORT)
async def leave_register(
    request: Request,
    from_date: str = Query(),
    to_date: str = Query(),
    department_code: str | None = Query(None),
    format: str = Query("xlsx", pattern="^(xlsx|pdf)$"),
    _: dict = Depends(require_role(*REPORT_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    rows = await _fetch_leave_register_rows(db, date.fromisoformat(from_date), date.fromisoformat(to_date), department_code)
    headers = ["Emp Code", "Name", "Dept", "Leave Type", "From", "To", "Days", "Status", "Approval Date"]
    body_rows = [
        [
            row["emp_code"],
            row["name"],
            row["dept"],
            row["leave_type"],
            row["from_date"],
            row["to_date"],
            row["days"],
            row["status"],
            row["approval_date"],
        ]
        for row in rows
    ]

    if format == "pdf":
        content = _pdf_bytes("Leave Register", headers, body_rows)
        return _stream_bytes(content, "application/pdf", f"leave_register_{from_date}_{to_date}.pdf")

    content = _workbook_bytes("Leave Register", headers, body_rows)
    return _stream_bytes(
        content,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"leave_register_{from_date}_{to_date}.xlsx",
    )


@router.get("/leave-abstract")
@limiter.limit(settings.RATE_LIMIT_EXPORT)
async def leave_abstract(
    request: Request,
    from_date: str = Query(),
    to_date: str = Query(),
    _: dict = Depends(require_role(*REPORT_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    rows = await _fetch_category_summary_rows(db, date.fromisoformat(from_date), date.fromisoformat(to_date))
    headers = ["Category", "Total Staff", "Total Leave Days by type", "Avg per staff"]
    body_rows = [[row["category"], row["total_staff"], row["total_leave_days_by_type"], row["avg_per_staff"]] for row in rows]
    content = _workbook_bytes("Category Summary", headers, body_rows)
    return _stream_bytes(
        content,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"category_leave_summary_{from_date}_{to_date}.xlsx",
    )


@router.get("/leave-abstract-department")
@limiter.limit(settings.RATE_LIMIT_EXPORT)
async def leave_abstract_department(
    request: Request,
    from_date: str = Query(),
    to_date: str = Query(),
    _: dict = Depends(require_role(*REPORT_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        text(
            """
            SELECT d.name AS department, lt.code AS leave_type, COUNT(*) AS count, SUM(a.applied_days) AS total_days
            FROM leave_applications a
            JOIN employees e ON a.employee_id = e.id
            JOIN departments d ON e.department_id = d.id
            JOIN leave_types lt ON a.leave_type_id = lt.id
            WHERE a.status = 'APPROVED' AND a.from_date BETWEEN :from_date AND :to_date
            GROUP BY d.name, lt.code
            ORDER BY d.name, lt.code
            """
        ),
        {"from_date": date.fromisoformat(from_date), "to_date": date.fromisoformat(to_date)},
    )
    return _rows_from_result(result)


@router.get("/pending-applications")
@limiter.limit(settings.RATE_LIMIT_EXPORT)
async def pending_applications(
    request: Request,
    _: dict = Depends(require_role(*REPORT_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    rows = await _fetch_pending_rows(db)
    headers = ["App #", "Employee", "Leave Type", "Submitted Date", "Days Pending", "Current Approver"]
    body_rows = [
        [
            row["app_number"],
            row["employee"],
            row["leave_type"],
            row["submitted_date"],
            row["days_pending"],
            row["current_approver"],
        ]
        for row in rows
    ]
    content = _pdf_bytes("Pending Applications (Aged)", headers, body_rows)
    return _stream_bytes(content, "application/pdf", "pending_applications_aged.pdf")


@router.get("/balance-summary")
@limiter.limit(settings.RATE_LIMIT_EXPORT)
async def balance_summary(
    request: Request,
    as_of_date: str | None = Query(None),
    department_code: str | None = Query(None),
    designation_name: str | None = Query(None),
    current_user: dict = Depends(require_role(*REPORT_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    rows = await _fetch_balance_rows(
        db,
        date.fromisoformat(as_of_date) if as_of_date else None,
        department_code,
        designation_name,
        current_user,
    )
    headers = ["Emp Code", "Name", "Dept", "Leave Type", "Opening", "Credited", "Availed", "Closing"]
    body_rows = [
        [
            row["emp_code"],
            row["name"],
            row["dept"],
            row["leave_type"],
            row["opening_balance"],
            row["credited"],
            row["availed"],
            row["closing_balance"],
        ]
        for row in rows
    ]
    content = _workbook_bytes("Balance Summary", headers, body_rows)
    suffix = as_of_date or str(date.today().year)
    return _stream_bytes(
        content,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"balance_summary_{suffix}.xlsx",
    )


@router.get("/balance-overview")
async def balance_overview(
    as_of_date: str | None = Query(None),
    department_code: str | None = Query(None),
    designation_name: str | None = Query(None),
    current_user: dict = Depends(require_role(*REPORT_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """Interactive balance grid for nodal officers — filter by department and designation."""
    rows = await _fetch_balance_rows(
        db,
        date.fromisoformat(as_of_date) if as_of_date else None,
        department_code,
        designation_name,
        current_user,
    )
    return {"as_of_date": as_of_date or str(date.today().year), "rows": rows, "count": len(rows)}


@router.get("/sanction-pdf/{application_id}")
@limiter.limit(settings.RATE_LIMIT_EXPORT)
async def sanction_pdf(
    request: Request,
    application_id: str,
    current_user: dict = Depends(require_role("ADMIN", "ESTABLISHMENT_OFFICER", "REGISTRAR", "DIRECTOR", "HOD", "DEAN_ACADEMIC")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        text(
            """
            SELECT a.*, e.name AS emp_name, e.emp_code,
                   d.name AS dept_name, des.name AS designation_name,
                   lt.name AS leave_type_name, lt.code AS leave_type_code
            FROM leave_applications a
            JOIN employees e ON a.employee_id = e.id
            JOIN departments d ON e.department_id = d.id
            JOIN designations des ON e.designation_id = des.id
            JOIN leave_types lt ON a.leave_type_id = lt.id
            WHERE a.id = :id
            """
        ),
        {"id": application_id},
    )
    app = result.fetchone()
    if not app:
        raise HTTPException(status_code=404)
    if app.status != "APPROVED":
        raise HTTPException(status_code=400, detail="Only APPROVED leave can generate sanction copy")

    buffer = io.BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm, topMargin=2 * cm, bottomMargin=2 * cm)
    story = []
    styles = getSampleStyleSheet()
    story.append(Paragraph("AIIMS Bibinagar", styles["Heading2"]))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph("<b>LEAVE SANCTION ORDER</b>", styles["Heading3"]))
    story.append(Spacer(1, 0.3 * cm))
    details = [
        ["Application No.", str(app.app_number)],
        ["Employee Name", f"{app.emp_name} ({app.emp_code})"],
        ["Department", app.dept_name],
        ["Designation", app.designation_name],
        ["Leave Type", f"{app.leave_type_name} ({app.leave_type_code})"],
        ["From Date", str(app.from_date)],
        ["To Date", str(app.to_date)],
        ["Applied Days", str(app.applied_days)],
        ["Status", "APPROVED"],
        ["Approving Authority", current_user.get("username", "-")],
        ["Date of Issue", str(date.today())],
    ]
    table = Table(details, colWidths=[5 * cm, 11 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#e8f0fe")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
            ]
        )
    )
    story.append(table)
    document.build(story)
    return _stream_bytes(buffer.getvalue(), "application/pdf", f"sanction_{str(app.app_number).replace('/', '-')}.pdf")


@router.get("/leave-calendar")
@limiter.limit(settings.RATE_LIMIT_EXPORT)
async def leave_calendar(
    request: Request,
    department_code: str | None = Query(None),
    month: str | None = Query(None),
    _: dict = Depends(require_role(*REPORT_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    rows = await _fetch_calendar_rows(db, department_code, month)
    headers = ["Emp Code", "Name", "From", "To", "Leave Type"]
    body_rows = [[row["emp_code"], row["name"], row["from_date"], row["to_date"], row["leave_type"]] for row in rows]
    content = _workbook_bytes("Leave Calendar", headers, body_rows)
    filename_suffix = month or "all"
    return _stream_bytes(
        content,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"leave_calendar_{filename_suffix}.xlsx",
    )


@router.get("/payroll-export")
@limiter.limit(settings.RATE_LIMIT_EXPORT)
async def payroll_export(
    request: Request,
    from_date: str = Query(),
    to_date: str = Query(),
    export_type: str = Query("LOP"),
    current_user: dict = Depends(require_role(*PAYROLL_EXPORT_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    if export_type != "LOP":
        raise HTTPException(status_code=400, detail="Only LOP export is supported in v1")

    fd = date.fromisoformat(from_date)
    td = date.fromisoformat(to_date)
    result = await db.execute(
        text(
            "SELECT "
            "e.emp_code AS emp_code, "
            "e.name AS name, "
            "d.name AS department, "
            "TO_CHAR(a.from_date, 'YYYY-MM') AS month, "
            "SUM(a.applied_days) AS lop_days, "
            "MIN(COALESCE(a.reason, '')) AS reason "
            "FROM leave_applications a "
            "JOIN employees e ON a.employee_id = e.id "
            "JOIN departments d ON e.department_id = d.id "
            "JOIN leave_types lt ON a.leave_type_id = lt.id "
            "WHERE a.from_date BETWEEN :from_date AND :to_date "
            "AND a.status = 'APPROVED' "
            "AND lt.code = 'EOL' "
            "GROUP BY e.emp_code, e.name, d.name, TO_CHAR(a.from_date, 'YYYY-MM') "
            "ORDER BY e.emp_code, month"
        ),
        {"from_date": fd, "to_date": td},
    )
    rows = _rows_from_result(result)

    await db.execute(
        text(
            "INSERT INTO payroll_export_log "
            "(id, export_from, export_to, export_type, exported_by, record_count, summary) "
            "VALUES "
            "(uuid_generate_v4(), :from_date, :to_date, :export_type, :exported_by, :record_count, CAST(:summary AS jsonb))"
        ),
        {
            "from_date": fd,
            "to_date": td,
            "export_type": export_type,
            "exported_by": current_user["user_id"],
            "record_count": len(rows),
            "summary": json.dumps(
                {
                    "mapping_note": "AIIMS Finance NIC spec pending; using placeholder export mapping",
                    "columns": list(PAYROLL_NIC_MAPPING_PLACEHOLDER.values()),
                }
            ),
        },
    )
    await db.commit()

    csv_buffer = io.StringIO()
    writer = csv.DictWriter(csv_buffer, fieldnames=list(PAYROLL_NIC_MAPPING_PLACEHOLDER.values()))
    writer.writeheader()
    for row in rows:
        writer.writerow(
            {
                PAYROLL_NIC_MAPPING_PLACEHOLDER["emp_code"]: row["emp_code"],
                PAYROLL_NIC_MAPPING_PLACEHOLDER["name"]: row["name"],
                PAYROLL_NIC_MAPPING_PLACEHOLDER["department"]: row["department"],
                PAYROLL_NIC_MAPPING_PLACEHOLDER["month"]: row["month"],
                PAYROLL_NIC_MAPPING_PLACEHOLDER["lop_days"]: row["lop_days"],
                PAYROLL_NIC_MAPPING_PLACEHOLDER["reason"]: row["reason"],
            }
        )
    content = csv_buffer.getvalue().encode("utf-8")
    return _stream_bytes(content, "text/csv; charset=utf-8", f"payroll_export_lop_{from_date}_{to_date}.csv")
